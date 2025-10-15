import streamlit as st
import pandas as pd
import re
from io import StringIO
from wordcloud import WordCloud
import matplotlib.pyplot as plt
import seaborn as sns
import nltk
from io import BytesIO
import numpy as np
from nltk.sentiment import SentimentIntensityAnalyzer
from collections import Counter
import plotly.express as px
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from matplotlib import font_manager
import networkx as nx

nltk.download('vader_lexicon')
HEADER_REGEX = re.compile(r"^(\d{1,2}/\d{1,2}/\d{2,4})\s*(上午|下午|早上|晚上)?\s*(\d{1,2}:\d{2})\s*-\s*(.*?):\s*(.*)$")
# Parse WhatsApp chat export txt into a dataframe
def parse_whatsapp(txt):
    pattern = r"(\d{1,2}/\d{1,2}/\d{2,4}) (晚上|早上)?(\d{1,2}:\d{2}) - (.*?): (.*)"
    rows = []
    for line in txt.split('\n'):
        m = re.match(pattern, line)
        if m:
            date_str = m.group(1)
            meridian = m.group(2)  # '晚上' or '早上' or None
            time_str = m.group(3)
            sender = m.group(4)
            message = m.group(5)
            # Convert to 24h time if needed
            hour, minute = map(int, time_str.split(':'))
            if meridian == '晚上' and hour < 12:
                hour += 12
            dt_str = f"{date_str} {hour}:{minute:02d}"
            dt = pd.to_datetime(dt_str, errors='coerce', dayfirst=False)
            if pd.notna(dt):
                rows.append({'datetime': dt, 'sender': sender, 'message': message})
    return pd.DataFrame(rows)

def parse_whatsapp_txt(text: str) -> pd.DataFrame:
    lines = text.splitlines()
    rows = []
    current = None
    for line in lines:
        line = line.strip()
        if not line:
            continue
        m = HEADER_REGEX.match(line)
        if m:
            # start of a new message
            date_str, ampm, time_str, sender, message = m.groups()
            # normalize am/pm if present -- many Chinese exports use 上午/下午 etc but hour is already in 24h sometimes
            # We'll try to parse using dayfirst
            try:
                dt = datetime.strptime(f"{date_str} {time_str}", "%d/%m/%Y %H:%M")
            except ValueError:
                # try 2-digit year
                try:
                    dt = datetime.strptime(f"{date_str} {time_str}", "%d/%m/%y %H:%M")
                except ValueError:
                    # fallback: parse with dayfirst=True via pandas
                    try:
                        dt = pd.to_datetime(f"{date_str} {time_str}", dayfirst=True)
                    except Exception:
                        dt = None
            if current:
                rows.append(current)
            current = {
                "Date": dt.date() if dt is not None else date_str,
                "Time": dt.time() if dt is not None else time_str,
                "message": message,
                "sender": sender,
                "datetime": dt,
            }
        else:
            # continuation of previous message
            if current is None:
                # orphan line (no header seen yet) - skip or treat as system
                continue
            current["message"] += "\n" + line

    if current:
        rows.append(current)

    df = pd.DataFrame(rows)
    # ensure datetime column is proper dtype
    if "datetime" in df.columns:
        df["datetime"] = pd.to_datetime(df["datetime"], errors="coerce")
    # keep ordering as in file
    return df

def clean_illegal_chars(text):
    if isinstance(text, str):
        # Removes control chars except \n, \r, \t (allowed in Excel)
        return re.sub(r'[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]', '', text)
    return text

def to_excel_bytes(df: pd.DataFrame) -> bytes:
    buffer = BytesIO()
    # write Excel with pandas
    df.to_excel(buffer, index=False, engine="openpyxl")
    buffer.seek(0)

    # Load workbook and worksheet
    wb = load_workbook(buffer)
    ws = wb.active

    # Set column widths (adjust columns as you want)
    widths = {'A': 10, 'B': 10, 'C': 15, 'D': 100}  # Example widths for Date, Time, sender, message
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for cell in ws['D']:
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    # Save workbook back to bytes buffer
    new_buffer = BytesIO()
    wb.save(new_buffer)
    new_buffer.seek(0)

    return new_buffer.read()


def compute_sender_cooccurrence(df: pd.DataFrame, k: int = 5) -> pd.DataFrame:
    """Compute co-occurrence matrix where value[current, previous] = count of times `previous` appeared within last k messages before `current`."""
    senders = list(df["sender"].astype(str))
    unique = sorted(list(set(senders)))
    idx = {s: i for i, s in enumerate(unique)}
    mat = np.zeros((len(unique), len(unique)), dtype=int) # rows: current, cols: previous
    for i, cur in enumerate(senders):
        cur_i = idx[cur]
        for j in range(1, k + 1):
            if i - j < 0:
                break
            prev = senders[i - j]
            prev_i = idx[prev]
            mat[cur_i, prev_i] += 1
    mat_df = pd.DataFrame(mat, index=unique, columns=unique)
    return mat_df

st.title("WhatsApp message History Analyzer")

uploaded = st.file_uploader("Upload WhatsApp message history text file (.txt)", type=['txt'])
use_stopwords = st.checkbox("Ignore common stopwords", value=True)
stopwords = set(["to", "a", "the", "and", "of", "in", "on", "for", "is", "it", "that", "with", "as", "at"])
if uploaded:
    txt = StringIO(uploaded.getvalue().decode('utf-8')).read()
    with st.spinner("Parsing file..."):
        df = parse_whatsapp(txt)
        df2 = parse_whatsapp_txt(txt)
    
    if df.empty:
        st.error("Could not parse the chat file. Make sure it’s a WhatsApp exported txt file.")
        st.stop()
        
    st.success(f"Loaded {len(df)} messages")

    #min_date = df['datetime'].dt.date.min()
    #max_date = df['datetime'].dt.date.max()
    #start_date, end_date = st.slider(
    #    "Select date range",
    #    min_value=min_date,
    #    max_value=max_date,
    #    value=(min_date, max_date)
    #)
    #df_filtered = df[(df['datetime'].dt.date >= start_date) & (df['datetime'].dt.date <= end_date)]
    #st.write(f"Showing {len(df_filtered)} messages between {start_date} and {end_date}")
    
    # slider in month, instead of in days
    min_month = df['datetime'].dt.to_period('M').min().to_timestamp()
    max_month = df['datetime'].dt.to_period('M').max().to_timestamp()

    months = pd.date_range(min_month, max_month, freq='MS').to_pydatetime().tolist()
    month_strs = [m.strftime('%Y-%m') for m in months]

    start_str, end_str = st.select_slider(
        "Select month range",
        options=month_strs,
        value=(month_strs[0], month_strs[-1])
    )

    start_date = pd.to_datetime(start_str + '-01')
    end_date = pd.to_datetime(end_str + '-01').replace(day=28) + pd.offsets.MonthEnd(0)

    df_filtered = df[(df['datetime'] >= start_date) & (df['datetime'] <= end_date)]

    st.write(f"Showing {len(df_filtered)} messages between {start_str} and {end_str}")

    tab1, tab2, tab3, tab4, tab5 = st.tabs(['Word Cloud', 'Top20 freq word', 'Monthly freq', 'Co-occurrence heatmap', 'Excel download'])
    with tab1:
        #1 Word Cloud
        #st.write(f"[1] Word Cloud")
        st.title("[1] Word Cloud")
        text = " ".join(df_filtered['message'].dropna())
        text = text.replace("媒體已略去", "")
        if use_stopwords:
            words = [w for w in re.findall(r'\b\w+\b', text.lower()) if w not in stopwords]
            text = " ".join(words)
        if text.strip():
            wc = WordCloud(font_path='NotoSansTC-Thin.ttf',width=1200, height=900, background_color='white').generate(text)
            fig, ax = plt.subplots(figsize=(10, 5))
            ax.imshow(wc, interpolation='bilinear')
            ax.axis('off')
            st.pyplot(fig)

        #2 Sentiment Intensity analysis
        #sia = SentimentIntensityAnalyzer()
        #df_filtered['sentiment'] = df_filtered['message'].apply(lambda x: sia.polarity_scores(x)['compound'])

        #sentiment_counts = df_filtered['sentiment'].apply(
        #    lambda s: 'Positive' if s>0.05 else ('Negative' if s<-0.05 else 'Neutral')
        #).value_counts()

        #st.bar_chart(sentiment_counts)
        #st.write(f"Avg sentiment: {df_filtered['sentiment'].mean():.3f}")

    with tab2:
        #2 Top 20 frequent words/phrases
        st.title("Top 20 Frequent words")
        # After filtering df_filtered with selected date range
        messages = df_filtered['message'].dropna().str.lower().str.replace("媒體已略去", "").tolist()

        # Simple tokenization: split by non-word chars, filter empty
        words = []
        for msg in messages:
            if use_stopwords:
                words.extend([w for w in re.findall(r'\b\w+\b', msg) if w not in stopwords])
            else:
                words.extend(re.findall(r'\b\w+\b', msg))

        # Count frequencies
        counter = Counter(words)
        top20 = counter.most_common(20)

        # Prepare data for plotting
        df_top20 = pd.DataFrame(top20, columns=['word', 'count'])

        # Plot horizontal bar chart with Plotly
        fig_freq = px.bar(df_top20.sort_values('count'), 
                        x='count', y='word',
                        orientation='h',
                        title='[2] Top 20 Frequent words',
                        labels={'word': 'Words', 'count': 'f'})
        fig_freq.update_layout(
        height=400,                            # taller figure to fit more words
        margin=dict(l=170, r=30, t=50, b=50),   # more left margin for long words
        yaxis=dict(tickfont=dict(size=10))      # smaller y-axis font size if needed
        )
        st.plotly_chart(fig_freq, use_container_width=True)

    with tab3:
        #3 Plot frequency of messages per month in selected period
        #st.subheader("[3] Monthly frequency")
        st.title("[3] Monthly frequency")
        #df_filtered
        df_filtered = df2[(df2['datetime'] >= start_date) & (df2['datetime'] <= end_date)]
        if df_filtered["datetime"].isna().all():
            st.warning("No valid datetime parsed, cannot compute monthly counts.")
        else:
            df_time = df_filtered.dropna(subset=["datetime"]).set_index("datetime").sort_index()
        # filter
        df_period = df_time.loc[start_date:end_date]
        monthly = df_period.resample('M').size().reset_index(name='count')
        fig = px.line(monthly, x='datetime', y='count', title='messages per month')
        fig.update_layout(xaxis_title='Month', yaxis_title='message count')
        st.plotly_chart(fig, use_container_width=True)

    with tab4:
        #4 Correlation / co-occurrence heatmap among senders based on previous k messages
        #st.subheader("[4] sender co-occurrence heatmap (previous k messages)")
        st.title("[4] sender co-occurrence heatmap (previous k messages)")
        
        k = st.slider("k (number of previous messages to consider)", min_value=1, max_value=10, value=5)

        if len(df2) == 0:
            st.write("No messages to analyze.")
        else:
            df_filtered = df2[(df2['datetime'] >= start_date) & (df2['datetime'] <= end_date)]
            co_mat = compute_sender_cooccurrence(df_period, k=k)    #df_filtered
            st.write("Raw counts (rows=current sender, cols=previous sender)")
            st.write("how often did sender X(col) talk just before sender Y(row)")
            st.markdown("""
                <style>.dataframe td,.dataframe th {
                    font-size: 7px!important;
                }
                </style>
            """, unsafe_allow_html=True)
            st.dataframe(co_mat)
            # Normalize rows to show relative frequency that a previous sender appears before current
            row_sums = co_mat.sum(axis=1).replace(0, 1)
            co_norm = co_mat.div(row_sums, axis=0)
            cmap = sns.light_palette("green", as_cmap=True)
            #font_files = font_manager.findSystemFonts(fontpaths='')
            #for font_file in font_files:
            #    font_manager.fontManager.addfont(font_file)
            font_manager.fontManager.addfont('NotoSansTC-Thin.ttf')
            font = font_manager.FontProperties(fname='NotoSansTC-Thin.ttf')
            plt.rcParams['font.sans-serif'] = [font.get_name(),'Microsoft JhengHei', 'Noto Sans CJK TC', 'Arial Unicode MS']
            plt.rcParams['axes.unicode_minus'] = False  # to show minus correctly
            fig2, ax2 = plt.subplots(figsize=(14, 10))
            sns.heatmap(co_norm, annot=True, fmt='.2f', ax=ax2, cmap=cmap, cbar_kws={'label': 'P(previous | current) normalized by row'},
                annot_kws={"size": 8})
            ax2.set_title(f"Normalized co-occurrence (previous up to {k} messages)")
            plt.xticks(rotation=45)
            plt.yticks(rotation=0)
            st.pyplot(fig2)

            # Optionally compute Pearson correlation between sender columns (based on co-occurrence vectors)
            if st.checkbox("Also show Pearson correlation of sender co-occurrence vectors"):
                # compute correlation matrix among rows of co_norm
                corr = co_norm.T.corr()
                fig3, ax3 = plt.subplots(figsize=(10, 8))
                sns.heatmap(corr, annot=True, fmt='.2f', ax=ax3)
                ax3.set_title("Pearson correlation between senders (based on co-occurrence patterns)")
                plt.xticks(rotation=45)
                plt.yticks(rotation=0)
                st.pyplot(fig3)

        #5 Visual flows and directed graph among current senders and previous senders
        st.subheader("[5] Visual flows and Directed graph")
        #senders = list(df_period["sender"].astype(str))
        #senders = co_mat.index.tolist()
        senders = co_norm.index.tolist()
        
        # Build directed graph from co-occurrence matrix (prev_sender -> current_sender)
        G = nx.DiGraph()
        for i, cur_sender in enumerate(senders):
            for j, prev_sender in enumerate(senders):
                weight = co_norm.iloc[i, j]
                #weight = co_mat.iloc[i, j]
                if weight > 0:
                    G.add_edge(prev_sender, cur_sender, weight=weight)

        # Detect clusters/communities (using simple connected components on undirected version)
        undirected_G = G.to_undirected()
        clusters = list(nx.connected_components(undirected_G))
        cluster_map = {}
        for i, cluster in enumerate(clusters):
            for node in cluster:
                cluster_map[node] = i

        # Set colors by cluster
        colors = ['lightblue', 'lightgreen', 'lightcoral', 'orange']
        node_colors = [colors[cluster_map[node] % len(colors)] for node in G.nodes()]

        #pos = nx.spring_layout(G, seed=42)  # layout with no overlap
        #pos = nx.kamada_kawai_layout(G)  # or try nx.circular_layout(G)
        pos = nx.circular_layout(G)
        plt.figure(figsize=(20,14))
        nx.draw_networkx_nodes(G, pos, node_size=3000, node_color=node_colors)
        edges = G.edges(data=True)
        weights = [d['weight'] for (u,v,d) in edges]
        nx.draw_networkx_edges(G, pos, arrowstyle='-|>', arrowsize=25, width=[w*3 for w in weights], connectionstyle='arc3, rad=0.1')
        nx.draw_networkx_labels(G, pos, font_size=16)

        plt.title("Chat flow and clusters between specific senders")
        plt.axis('off')
        st.pyplot(plt.gcf())

    with tab5:
        #6 Re-format whatsapp message history in Excel for download
        # Offer Excel download (Date in col A, Time in col B, message in col C as the user requested). We'll order columns accordingly and include sender as extra column.
        #st.subheader(f"[6] Re-format whatsapp message history for Excel output")
        st.title(f"[6] Re-format whatsapp message history for Excel output")
        export_df = df2[(df2['datetime'] >= start_date) & (df2['datetime'] <= end_date)]
        # Create columns in the requested order
        export_df_for_excel = export_df[["Date", "Time", "sender", "message", ]]
        for col in ["Date", "Time", "sender", "message"]:
            export_df_for_excel[col] = export_df_for_excel[col].apply(clean_illegal_chars)

        excel_bytes = to_excel_bytes(export_df_for_excel)
        st.download_button("Download parsed messages as Excel", data=excel_bytes, file_name="whatsapp_("+uploaded.name+").xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

else:
    st.info("Upload a WhatsApp chat.txt file to start analyzing.")

